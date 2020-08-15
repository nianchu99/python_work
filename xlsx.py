import openpyxl, datetime

wb = openpyxl.load_workbook('Test.xlsx')

print(wb.sheetnames) # 这里既不能使用get_sheet_names()也不能使用sheetnames(),只能使用sheetnames
'''
print(wb.get_sheet_by_name('Sheet1')) # 提示不建议wb.get_sheet_by_name('Sheet1')这种方法'''
print(wb['Sheet1'])
print(wb['Sheet1'].title)
anotherSheet = wb.active # 获得当前活动表
print(anotherSheet.title) # 打印当前活动表 题目
sheet = wb['Sheet1']
value = sheet['A1'].value
print(value)

c = sheet['B1'] # row是A1中的1，column是其中的A
print(str(c.row) + str(c.column) + " is " + str(c.value))
sheet['B1'] = 'leibowen';
# for循环输出某一列的值
for i in range(1, 8):
    string = "B" + str(i)
    c = sheet[string]
    print(c.value)
# 使用column和row的组合来代替 A1，A2等
print("——————————————————————————————————————————")
for row in range(1,8,2):# 后一个数字是步长
    print(sheet.cell(row = row, column = 2).value)

# 确定表的大小
'''
highColumn = sheet.get_highest_column()
highRow = sheet.get_highest_raw()
print(str(highColumn) + str(highRow))
'''
print(openpyxl.get_column_letter(27))

